import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

#Вспомогательные функции для работы с файлом
def find_index(a,sheet,count):
    index = 0
    sheet_shablon = sheet
    row_count = count
    for i in range(1, row_count + 1):
        if sheet_shablon.cell(row=i, column=1).value == a:
            index = sheet_shablon.cell(row=i, column=1).row  # нашли строку от которой нужно выбирать данные
            break
    return index
def open_wb(file):
    wb = load_workbook(file.file)
    return wb
def give_name_file(file):
    name_file = file.file.name[6:][:-5]
    return name_file

#Функции проверки без изменения
def give_list(i1, j1, sheet, sheet2,m,n):
    list = []
    sheet_shablon = sheet
    sheet_wb = sheet2
    a=m
    b=n
    for i in range(a + 2, b):
        for j in range(i1, j1):
            list.append(sheet_shablon.cell(row=i, column=j).value)
    k = 0
    for i in range(a + 2, b):
        for j in range(i1, j1):
            if (sheet_wb.cell(row=i, column=j).value) != list[k]:
                sheet_wb.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
                k = k + 1
            else:
                sheet_wb.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
                k = k + 1
    return locals()
def check_titul(req, file,shab):
    wb = open_wb(file)
    sheet = wb.get_sheet_by_name('Титул')
    shablon = load_workbook(shab)
    sheet_shablon = shablon['Титул']
    name_file = give_name_file(file)
    napravlenie = ''
    if name_file[:1] == '0':
        napravlenie = name_file[:8]
    if name_file[:1] == 'З':
        napravlenie = name_file[2:10]
    if name_file[:3] == 'О-З':
        napravlenie = name_file[4:12]
    god_f = name_file[-4:]
    if sheet['B16'].value == napravlenie:
        sheet['B16'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    else:
        sheet['B16'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    if sheet['T29'].value == god_f:
        sheet['T29'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    else:
        sheet['T29'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    if sheet['B10'].value != sheet_shablon['B10'].value:
        sheet['B10'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['B10'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['B18'].value != sheet_shablon['B18'].value:
        sheet['B18'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['B18'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['B19'].value != sheet_shablon['B19'].value:
        sheet['B19'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['B19'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['B26'].value != sheet_shablon['B26'].value:
        sheet['B26'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['B26'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['R30'].value != sheet_shablon['R30'].value:
        sheet['R30'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['R30'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['R31'].value != sheet_shablon['R31'].value:
        sheet['R31'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['R31'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A29'].value != sheet_shablon['A29'].value:
        sheet['A29'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['A29'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A30'].value != sheet_shablon['A30'].value:
        sheet['A30'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['A30'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A31'].value != sheet_shablon['A31'].value:
        sheet['A31'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['A31'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A32'].value != sheet_shablon['A32'].value:
        sheet['A32'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['A32'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['W38'].value != sheet_shablon['W38'].value:
        sheet['W38'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['W38'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['W40'].value != sheet_shablon['W40'].value:
        sheet['W40'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['W40'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['W42'].value != sheet_shablon['W42'].value:
        sheet['W42'].fill = PatternFill(fill_type='solid', start_color='fe0101', end_color='fe0101')
    else:
        sheet['W42'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def check_grafik(req,file,shab):
    wb = open_wb(file)
    sheet_wb = wb['График']
    shablon = load_workbook(shab)
    sheet_shablon = shablon['График']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='I', sheet = sheet_shablon, count=row_count) - 2
    b = find_index(a='IV', sheet = sheet_shablon, count=row_count) + 6
    list = give_list(2, 54, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def check_plan(req,file,shab):
    wb = open_wb(file)
    shablon = load_workbook(shab)
    sheet_wb = wb['План']
    sheet_shablon = shablon['План']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='Блок 1.Дисциплины (модули) ',sheet = sheet_shablon, count=row_count)
    b = find_index(a='Вариативная часть ',sheet = sheet_shablon, count=row_count)
    list = give_list(2, 69, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    # Блок2 практики
    a = find_index(a='Блок 2.Практики ',sheet = sheet_shablon, count=row_count)
    b = find_index(a='Блок 3.Государственная итоговая аттестация ',sheet = sheet_shablon, count=row_count)
    list = give_list(2, 69, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def check_plan_svod(req,file,shab):
    wb = open_wb(file)
    shablon = load_workbook(shab)
    sheet_wb = wb['ПланСвод']
    sheet_shablon = shablon['ПланСвод']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='Блок 1.Дисциплины (модули) ',sheet = sheet_shablon, count=row_count)
    b = find_index(a='Вариативная часть ',sheet = sheet_shablon, count=row_count)
    list = give_list(2, 26,sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def check_compititions(req,file,shab):
    wb = open_wb(file)
    shablon = load_workbook(shab)
    sheet_wb = wb['Компетенции(2)']
    sheet_shablon = shablon['Компетенции(2)']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='Б1', sheet = sheet_shablon, count=row_count)
    b = find_index(a='ФТД', sheet = sheet_shablon, count=row_count)
    list = give_list(3, 7, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def check_all(req,file, shab):
    check_titul(req, file, shab)
    check_grafik(req, file, shab)
    check_plan(req, file, shab)
    check_plan_svod(req, file, shab)
    check_compititions(req, file, shab)
    return locals()

#Функции проверки с изменением
def edit_list(i1, j1, sheet, sheet2,m,n):
    list = []
    sheet_shablon = sheet
    sheet_wb = sheet2
    a=m
    b=n
    for i in range(a + 2, b):
        for j in range(i1, j1):
            list.append(sheet_shablon.cell(row=i, column=j).value)
    k = 0
    for i in range(a + 2, b):
        for j in range(i1, j1):
            if (sheet_wb.cell(row=i, column=j).value) != list[k]:
                sheet_wb.cell(row=i, column=j).value = list[k]
                sheet_wb.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='38a0f5',
                                                                  end_color='38a0f5')
                k = k + 1
            else:
                sheet_wb.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
                k = k + 1
    return locals()
def edit_titul(req, file,shab):
    wb = open_wb(file)
    sheet = wb.get_sheet_by_name('Титул')
    shablon = load_workbook(shab)
    sheet_shablon = shablon['Титул']
    name_file = give_name_file(file)
    napravlenie = ''
    if name_file[:1] == '0':
        napravlenie = name_file[:8]
    if name_file[:1] == 'З':
        napravlenie = name_file[2:10]
    if name_file[:3] == 'О-З':
        napravlenie = name_file[4:12]
    god_f = name_file[-4:]
    if sheet['B16'].value == napravlenie and sheet['T29'].value == god_f:
        sheet['B16'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
        sheet['T29'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    else:
        sheet['B16'].value = napravlenie
        sheet['T29'].value = god_f
        sheet['B16'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
        sheet['T29'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    if sheet['B10'].value != sheet_shablon['B10'].value:
        sheet['B10'].value = sheet_shablon['B10'].value
        sheet['B10'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['B10'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['B18'].value != sheet_shablon['B18'].value:
        sheet['B18'].value = sheet_shablon['B18'].value
        sheet['B18'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['B18'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['B19'].value != sheet_shablon['B19'].value:
        sheet['B19'].value = sheet_shablon['B19'].value
        sheet['B19'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['B19'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['B26'].value != sheet_shablon['B26'].value:
        sheet['B26'].value = sheet_shablon['B26'].value
        sheet['B26'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['B26'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['R30'].value != sheet_shablon['R30'].value:
        sheet['R30'].value = sheet_shablon['R30'].value
        sheet['R30'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['R30'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['R31'].value != sheet_shablon['R31'].value:
        sheet['R31'].value = sheet_shablon['R31'].value
        sheet['R31'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['R31'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A29'].value != sheet_shablon['A29'].value:
        sheet['A29'].value = sheet_shablon['A29'].value
        sheet['A29'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['A29'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A30'].value != sheet_shablon['A30'].value:
        sheet['A30'].value = sheet_shablon['A30'].value
        sheet['A30'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['A30'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A31'].value != sheet_shablon['A31'].value:
        sheet['A31'].value = sheet_shablon['A31'].value
        sheet['A31'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['A31'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['A32'].value != sheet_shablon['A32'].value:
        sheet['A32'].value = sheet_shablon['A32'].value
        sheet['A32'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['A32'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['W38'].value != sheet_shablon['W38'].value:
        sheet['W38'].value = sheet_shablon['W38'].value
        sheet['W38'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['W38'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['W40'].value != sheet_shablon['W40'].value:
        sheet['W40'].value = sheet_shablon['W40'].value
        sheet['W40'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['W40'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    if sheet['W42'].value != sheet_shablon['W42'].value:
        sheet['W42'].value = sheet_shablon['W42'].value
        sheet['W42'].fill = PatternFill(fill_type='solid', start_color='38a0f5', end_color='38a0f5')
    else:
        sheet['W42'].fill = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def edit_grafik(req,file,shab):
    wb = open_wb(file)
    sheet_wb = wb['График']
    shablon = load_workbook(shab)
    sheet_shablon = shablon['График']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='I', sheet = sheet_shablon, count=row_count) - 2
    b = find_index(a='IV', sheet = sheet_shablon, count=row_count) + 6
    list = edit_list(2, 54, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def edit_plan(req,file,shab):
    wb = open_wb(file)
    shablon = load_workbook(shab)
    sheet_wb = wb['План']
    sheet_shablon = shablon['План']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='Блок 1.Дисциплины (модули) ',sheet = sheet_shablon, count=row_count)
    b = find_index(a='Вариативная часть ',sheet = sheet_shablon, count=row_count)
    list = give_list(2, 69, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    # Блок2 практики
    a = find_index(a='Блок 2.Практики ',sheet = sheet_shablon, count=row_count)
    b = find_index(a='Блок 3.Государственная итоговая аттестация ',sheet = sheet_shablon, count=row_count)
    list = edit_list(2, 69, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def edit_plan_svod(req,file,shab):
    wb = open_wb(file)
    shablon = load_workbook(shab)
    sheet_wb = wb['ПланСвод']
    sheet_shablon = shablon['ПланСвод']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='Блок 1.Дисциплины (модули) ',sheet = sheet_shablon, count=row_count)
    b = find_index(a='Вариативная часть ',sheet = sheet_shablon, count=row_count)
    list = edit_list(2, 26,sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def edit_compititions(req,file,shab):
    wb = open_wb(file)
    shablon = load_workbook(shab)
    sheet_wb = wb['Компетенции(2)']
    sheet_shablon = shablon['Компетенции(2)']
    name_file = give_name_file(file)
    row_count = sheet_shablon.max_row
    a = find_index(a='Б1', sheet = sheet_shablon, count=row_count)
    b = find_index(a='ФТД', sheet = sheet_shablon, count=row_count)
    list = edit_list(3, 7, sheet = sheet_shablon, sheet2=sheet_wb, m=a, n=b)
    wb.save('media/' + 'media/' + name_file + '.xlsx')
    wb.close()
    return locals()
def edit_all(req,file, shab):
    edit_titul(req, file, shab)
    edit_grafik(req, file, shab)
    edit_plan(req, file, shab)
    edit_plan_svod(req, file, shab)
    edit_compititions(req, file, shab)
    return locals()

#Функция проверки формата файла
def valid_file(value):
    from django.core.exceptions import ValidationError
    if not value.name.endswith('.xlsx'):
        raise ValidationError('Неподдерживаемый формат файла')

